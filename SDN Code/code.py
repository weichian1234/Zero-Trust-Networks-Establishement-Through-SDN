from pox.core import core
import pox.openflow.libopenflow_01 as of
from pox.lib.util import dpidToStr
from pox.lib.revent import *
from pox.lib.addresses import IPAddr, EthAddr
import pox.lib.packet as pkt

import csv
import os
import pandas as pd
from openpyxl import load_workbook

log = core.getLogger()
policyFile = "%s/pox/ext/firewallPolicy.csv" % os.environ[ 'HOME' ] # One excel file for all rules
ruleFile = "%s/pox/ext/firewallRule.xlsx" % os.environ[ 'HOME' ] #ignore
host1Policy = "%s/pox/ext/host1lvl2Policy.csv" % os.environ[ 'HOME' ] #ignore
host3Policy = "%s/pox/ext/host3lvl2Policy.csv" % os.environ[ 'HOME' ] #ignore
host5Policy = "%s/pox/ext/host5lvl2Policy.csv" % os.environ[ 'HOME' ] #ignore
host7Policy = "%s/pox/ext/host7lvl2Policy.csv" % os.environ[ 'HOME' ] #ignore

class LearningSwitch(object):
     def __init__(self, connection):
         self.connection = connection
         connection.addListeners(self)
         self.list1 = [] #ignore
         self.host1 = [] #ignore
         self.host3 = [] #ignore
         self.host5 = [] #ignore
         self.host7 = [] #ignore
         self.priveleges = {}  #ignore
         
     def readRule(self):
         print("Before read Excel File")
         rule = pd.ExcelFile(ruleFile)
         print("After read Excel File")
         self.df = rule.parse("basic")
         new = self.df.set_index('Ip_Address').T.to_dict()
         df2 = rule.parse("privileges")
         self.priveleges = df2.set_index('User').T.to_dict()
         df3 = rule.parse("department")
         self.department = df3.set_index('Department').T.to_dict()
         self.wb = load_workbook(ruleFile)
         self.sheet1 = self.wb.get_sheet_by_name("basic")
         #self.storeIp = self.df.groupby('Switch_ID')['Ip_Address'].apply(list)
         #print (new)
         return new
         
     def get_cell(self, src_ip):
         srcip = str(src_ip)
         row_num = self.df.index[self.df.Ip_Address == srcip][0]
         self.cell = self.sheet1.cell(row=row_num+2, column=5)
         self.cell1 = self.sheet1.cell(row=row_num+2, column=4)
         print('Cell value: ', self.cell.value)
         self.cell.value = self.cell.value - 10
         self.check_value(self.cell.value)
     
     def check_value(self, value_parse):
         print ("user Value", self.cell1.value)
         print ("lvl Value", type(self.cell.value))
         
         if self.cell1.value == "default":
             return
         if value_parse <40:
             self.cell1.value = "default"
             self.cell.value = 100
             self.wb.save(ruleFile)
             
     
     def readlvl2(self, filepath): #ignore
         ifile = open(filepath,"r")
         if(ifile):
             print ("Open the lvl2 file successfully")
         else:
             print ("Filed to open lvl2 file")
         reader = csv.reader(ifile)
         rownum = 0
         l = []
         for row in reader:
                 if rownum == 0:
                      rownum += 1
                      print ("Continue")
                      continue                     
                 else:
                      print ("Check the Rule")
                      for col in row:
                          l.append(col)
                      
                 #l = []
                 rownum = rownum + 1
         print ("The list yeoh : ", l)
         ifile.close()
         return l

     def writeFile(self): #ignore
         
         ifile = open(policyFile,"r")
         if(ifile):
             print ("Open the file successfully")
         else:
             print ("GG liao la")
         reader = csv.reader(ifile)         

         rownum = 0
         l = []
         for row in reader:
                 if rownum == 0:
                      rownum += 1
                      print ("Continue")
                      continue                     
                 else:
                      print ("Check the Rule")
                      for col in row:
                          l.append(col)
                      
                 #l = []
                 rownum = rownum + 1
         #print ("The list yeoh : ", l)
         self.list1 = l
         #print ("The list One yeoh : ", self.list1)
         l= []
         ifile.close()
          
     def writeRule(self, packet, port):
         print ("Writing rule.....")
         msg = of.ofp_flow_mod()
         msg.match = of.ofp_match.from_packet(packet, port)     
         msg.idle_timeout = 10
         msg.hard_timeout = 10
         msg.command = of.OFPFC_MODIFY     
         msg.priority = 11
         self.connection.send(msg)
    
     def allowRule(self, src, dst): #ignore
         print ("Writing rule.....")
         msg = of.ofp_flow_mod()
         match = of.ofp_match(dl_type = 0x800, nw_proto= pkt.ipv4.ICMP_PROTOCOL)
         match.nw_src = IPAddr(src)
         match.nw_dst = IPAddr(dst)
         msg.match = match
         msg.idle_timeout = 10
         msg.hard_timeout = 10
         msg.actions.append(of.ofp_action_output(port = of.OFPP_NORMAL))      
         msg.priority = 10
         self.connection.send(msg)

     def checkToken(self, srcip):
         row_num = self.df.index[self.df.Ip_Address == srcip][0]
         self.token = self.sheet1.cell(row=row_num+2, column=7)
         self.token_time = self.sheet1.cell(row=row_num+2, column=8)
         self.token_time.value = self.token_time.value - 1
         if self.token_time.value == 0:
         	self.token.value = 'no'
         self.wb.save(ruleFile)
         
     def checkTCP(self,pass_rule, src_ip, dst_ip, pass_packet):
         print('Check TCP')
         tcp = pass_packet.find('tcp')
         check_ip = str(src_ip)
         if pass_packet.find('tcp') is not None:
             dstport = str(tcp.dstport)
             if tcp.dstport == 20:
             	return False
             	
             if tcp.dstport in range (self.department['Manager']['Start'],self.department['Manager']['End']+1) or tcp.srcport in range (self.department['Manager']['Start'],self.department['Manager']['End']+1):
             	return False
             
             if tcp.dstport == self.priveleges[pass_rule[check_ip]['Current_User']]['userID']:
             	print ('User got port')
             	if self.priveleges[pass_rule[check_ip]['Current_User']]['Dep'] == pass_rule[dst_ip]['Device_dep']:
             		print ('User within department')
             		print('User IP',self.priveleges[pass_rule[check_ip]['Current_User']]['Dep'])
             		print('Device IP',pass_rule[check_ip]['Device_dep'])
             		return False
             	if pass_rule[check_ip]['Token_time'] != 0 and pass_rule[check_ip]['Access_token'] != 'no':
             		print ('User not within department')
             		if pass_rule[check_ip]['Access_token'] == pass_rule[dst_ip]['Device_dep']:
             			print('Token Check')
             			self.checkToken(check_ip)
             			print('Token Check -return false')
             			return False
             print('Check TCP - Inner True')
             return True
             #print ('Source Port Number is ', tcp.srcport)
             #print ('Destination Port Number is ', tcp.dstport)
         else:
             print('Check TCP - Outer False')
             return False
         pass   
  
     def switchBehave(self, packet, packet_in, port_switch, switch_id): #consist of main functionality
         flowMsg = of.ofp_flow_mod()
         flowMsg.idle_timeout = 10
         flowMsg.hard_timeout = 10
         flowMsg.priority = 13
         flowMsg.match = of.ofp_match.from_packet(packet)
         flowMsg.data = packet_in
         
         #self.writeFile()
         rule = self.readRule()
         print (rule)
         print (rule['10.0.0.1']['lvl1'])
         self.host1 = self.readlvl2(host1Policy)
         self.host3 = self.readlvl2(host3Policy)
         self.host5 = self.readlvl2(host5Policy)
         self.host7 = self.readlvl2(host7Policy)
         print('Priveleges: ',self.priveleges)
         
         
   
         #packet type Boolean vars
         ip = packet.find('ipv4')
         icmp = packet.find('icmp')
         print("Current switch_id", switch_id)
         tcp = packet.find('tcp')
         
         if packet.find('tcp') is not None:
             print ('Source Port Number is ', tcp.srcport)
             print ('Destination Port Number is ', tcp.dstport)
         else:
             print ('no tcp')
         
               
         if packet.find('icmp') is not None:
             print ('I found ICMP packet')  
         if ip is None:
             msgAction = of.ofp_action_output(port = of.OFPP_FLOOD)
             flowMsg.actions.append(msgAction)

         else:
             print (ip.srcip, ip.dstip)
             src_ip = str(ip.srcip)
             dst_ip =str(ip.dstip)
             if rule[src_ip]['lvl1'] == 'no':
             	self.connection.send(flowMsg)
             	return
             if dst_ip in [rule[src_ip]['Prohibited_IP']]:
             	print('Yes, triger h2 ')
             	self.writeRule(packet,port_switch)
             	return
             if self.checkTCP(rule, ip.srcip, dst_ip, packet):
             	self.get_cell(ip.srcip)
             	self.wb.save(ruleFile)
             	self.writeRule(packet, port_switch)
             	return
             	
         msgAction = of.ofp_action_output(port = of.OFPP_NORMAL)
         flowMsg.actions.append(msgAction)
         self.connection.send(flowMsg)
         return
	 
     def _handle_PacketIn (self, event):
         packet = event.parsed
         if not packet.parsed:
             log.warning("Ignoring incomplete packet")
             return
  
         packet_in = event.ofp
         self.switchBehave(packet, packet_in, event.port,event.dpid)


def launch():
     def _handle_ConnectionUp (event):
         LearningSwitch(event.connection)

     core.openflow.addListenerByName("ConnectionUp", _handle_ConnectionUp)